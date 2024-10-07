import os
import sys
import uuid
from pathlib import Path

import numpy as np
import pandas as pd
import torch
from tqdm import tqdm

sys.path.append('./models/backbone')
sys.path.append('/kaggle/working/MuSc')

import datasets.mvtec as mvtec
from datasets.mvtec import _CLASSNAMES as _CLASSNAMES_mvtec_ad
import datasets.visa as visa
from datasets.visa import _CLASSNAMES as _CLASSNAMES_visa
import datasets.btad as btad
from datasets.btad import _CLASSNAMES as _CLASSNAMES_btad
import datasets.rayan_industrial as rayan_industrial
from datasets.rayan_industrial import _CLASSNAMES as _CLASSNAMES_rayan_industrial

from utils.metrics import compute_metrics
from utils.helpers import dict_to_json
from openpyxl import Workbook
import cv2

import warnings

warnings.filterwarnings("ignore")


class BaseEval:
    def __init__(self, cfg, seed=0):
        self.cfg = cfg
        self.seed = seed
        self.device = torch.device("cuda:{}".format(cfg['device']) if torch.cuda.is_available() else "cpu")

        self.path = cfg['datasets']['data_path']
        self.dataset = cfg['datasets']['dataset_name']
        self.vis = cfg['testing']['vis']
        self.vis_type = cfg['testing']['vis_type']
        self.save_excel = cfg['testing']['save_excel']
        self.should_save_scores = cfg['testing']['should_save_scores']
        # the categories to be tested
        self.categories = cfg['datasets']['class_name']
        if isinstance(self.categories, str):
            if self.categories.lower() == 'all':
                if self.dataset == 'visa':
                    self.categories = _CLASSNAMES_visa
                elif self.dataset == 'mvtec_ad':
                    self.categories = _CLASSNAMES_mvtec_ad
                elif self.dataset == 'btad':
                    self.categories = _CLASSNAMES_btad
                elif self.dataset == 'rayan_industrial':
                    self.categories = _CLASSNAMES_rayan_industrial
            else:
                self.categories = [self.categories]

        self.image_size = cfg['datasets']['img_resize']
        self.divide_num = cfg['datasets']['divide_num']

        self.model_name = cfg['models']['backbone_name']
        run_name = cfg['testing'].get('run_name', str(uuid.uuid4()))
        print(f'Run Name: {run_name}')
        self.output_dir = os.path.join(
            cfg['testing']['output_dir'],
            self.dataset,
            self.model_name,
            'imagesize{}'.format(self.image_size),
            run_name
        )
        self.scores_dir = os.path.join(self.output_dir, 'scores')
        os.makedirs(self.output_dir, exist_ok=True)

        self.save_scores_precision = cfg['testing']['save_scores_precision']


    def load_datasets(self, category, divide_num=1, divide_iter=0):
        dataset_classes = {
            'visa': visa.VisaDataset,
            'mvtec_ad': mvtec.MVTecDataset,
            'btad': btad.BTADDataset,
            'rayan_industrial': rayan_industrial.RayanIndustrialDataset
        }

        dataset_splits = {
            'visa': visa.DatasetSplit.TEST,
            'mvtec_ad': mvtec.DatasetSplit.TEST,
            'btad': btad.DatasetSplit.TEST,
            'rayan_industrial': mvtec.DatasetSplit.TEST
        }

        test_dataset = dataset_classes[self.dataset](
            source=self.path,
            split=dataset_splits[self.dataset],
            classname=category,
            resize=self.image_size,
            imagesize=self.image_size,
            clip_transformer=self.preprocess,
            divide_num=divide_num,
            divide_iter=divide_iter,
            random_seed=self.seed
        )
        return test_dataset

    def visualization(self, image_path_list, gt_list, pr_px, category):
        def normalization01(img):
            return (img - img.min()) / (img.max() - img.min())

        if self.vis_type == 'single_norm':
            # normalized per image
            for i, path in enumerate(image_path_list):
                anomaly_type = path.split('/')[-2]
                img_name = path.split('/')[-1]
                if anomaly_type not in ['good', 'Normal', 'ok'] and gt_list[i] != 0:
                    save_path = os.path.join(self.output_dir, category, anomaly_type)
                    os.makedirs(save_path, exist_ok=True)
                    save_path = os.path.join(save_path, img_name)
                    anomaly_map = pr_px[i].squeeze()
                    anomaly_map = normalization01(anomaly_map) * 255
                    anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)
                    cv2.imwrite(save_path, anomaly_map)
        else:
            # normalized all image
            pr_px = normalization01(pr_px)
            for i, path in enumerate(image_path_list):
                anomaly_type = path.split('/')[-2]
                img_name = path.split('/')[-1]
                save_path = os.path.join(self.output_dir, category, anomaly_type)
                os.makedirs(save_path, exist_ok=True)
                save_path = os.path.join(save_path, img_name)
                anomaly_map = pr_px[i].squeeze()
                anomaly_map *= 255
                anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)
                cv2.imwrite(save_path, anomaly_map)

    def make_category_data(self, category):
        print(category)

        gt_sp, pr_sp, gt_px, pr_px, image_path_list = self.get_category_scores(category)
        if self.should_save_scores:
            self.save_scores(image_path_list, pr_sp, pr_px)

        pr_px, image_metric, pixel_metric = self.compute_category_metrics(category, gt_sp, pr_sp, gt_px, pr_px)

        if self.vis:
            print('visualization...')
            self.visualization(image_path_list, gt_sp.tolist(), pr_px, category)

        return image_metric, pixel_metric

    def get_category_scores(self, category):
        raise NotImplementedError()

    def compute_category_metrics(self, category, gt_sp, pr_sp, gt_px, pr_px):
        """
            pr_sp: Predicted Scores for Image-Level Classification | shape: (num_classes, )
            gt_sp: Ground Truth for Image-Level Classification | shape: (num_classes, )

            pr_px: Predicted Anomaly Maps for Pixel-Level Classification | shape: (num_classes, 1, width, height)
            gt_px: Ground Truth for Pixel-Level Classification | shape: (num_classes, 1, width, height)
        """
        print('computing metrics...')
        image_metric, pixel_metric = compute_metrics(gt_sp, pr_sp, gt_px, pr_px)

        auroc_sp, f1_sp, ap_sp = image_metric
        auroc_px, f1_px, ap_px, aupro = pixel_metric

        print(category)
        print('image-level, auroc:{}, f1:{}, ap:{}'.format(
            auroc_sp * 100, f1_sp * 100, ap_sp * 100)
        )
        print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(
            auroc_px * 100, f1_px * 100, ap_px * 100, aupro * 100)
        )
        return pr_px, image_metric, pixel_metric

    def save_scores(self, image_path_list, pr_sp, pr_px):
        print(f'Saving scores at {self.scores_dir}')
        for i in tqdm(range(len(image_path_list)), desc=f'Saving scores'):
            image_path = image_path_list[i]
            image_score_path = self.get_scores_path_for_image(image_path)
            os.makedirs(os.path.dirname(image_score_path), exist_ok=True)

            d = {
                'img_level_score': np.around(pr_sp[i], self.save_scores_precision),
                'pix_level_score': np.around(pr_px[i], self.save_scores_precision),
            }
            dict_to_json(d, image_score_path)

    def get_scores_path_for_image(self, image_path):
        # example image_path: './data/mvtec_anomaly_detection/cable/test/good/039.png'
        path = Path(image_path)

        category, split, anomaly_type = path.parts[-4:-1]
        image_name = path.stem

        return os.path.join(
            self.scores_dir, category, split, anomaly_type, f'{image_name}_scores.json'
        )

    def main(self):
        auroc_sp_ls = []
        f1_sp_ls = []
        ap_sp_ls = []
        auroc_px_ls = []
        f1_px_ls = []
        ap_px_ls = []
        aupro_ls = []
        for category in self.categories:
            image_metric, pixel_metric = self.make_category_data(category=category, )
            auroc_sp, f1_sp, ap_sp = image_metric
            auroc_px, f1_px, ap_px, aupro = pixel_metric
            auroc_sp_ls.append(auroc_sp)
            f1_sp_ls.append(f1_sp)
            ap_sp_ls.append(ap_sp)
            auroc_px_ls.append(auroc_px)
            f1_px_ls.append(f1_px)
            ap_px_ls.append(ap_px)
            aupro_ls.append(aupro)
        # mean
        auroc_sp_mean = sum(auroc_sp_ls) / len(auroc_sp_ls)
        f1_sp_mean = sum(f1_sp_ls) / len(f1_sp_ls)
        ap_sp_mean = sum(ap_sp_ls) / len(ap_sp_ls)
        auroc_px_mean = sum(auroc_px_ls) / len(auroc_px_ls)
        f1_px_mean = sum(f1_px_ls) / len(f1_px_ls)
        ap_px_mean = sum(ap_px_ls) / len(ap_px_ls)
        aupro_mean = sum(aupro_ls) / len(aupro_ls)

        for i, category in enumerate(self.categories):
            print(category)
            print('image-level, auroc:{}, f1:{}, ap:{}'.format(
                auroc_sp_ls[i] * 100, f1_sp_ls[i] * 100, ap_sp_ls[i] * 100)
            )
            print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(
                auroc_px_ls[i] * 100, f1_px_ls[i] * 100, ap_px_ls[i] * 100, aupro_ls[i] * 100)
            )

        print('mean')
        print('image-level, auroc:{}, f1:{}, ap:{}'.format(
            auroc_sp_mean * 100, f1_sp_mean * 100, ap_sp_mean * 100)
        )
        print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(
            auroc_px_mean * 100, f1_px_mean * 100, ap_px_mean * 100, aupro_mean * 100)
        )

        # save in excel
        if self.save_excel:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "MuSc_results"
            sheet.cell(row=1, column=2, value='auroc_px')
            sheet.cell(row=1, column=3, value='f1_px')
            sheet.cell(row=1, column=4, value='ap_px')
            sheet.cell(row=1, column=5, value='aupro')
            sheet.cell(row=1, column=6, value='auroc_sp')
            sheet.cell(row=1, column=7, value='f1_sp')
            sheet.cell(row=1, column=8, value='ap_sp')
            for col_index in range(2):
                for row_index in range(len(self.categories)):
                    if col_index == 0:
                        sheet.cell(row=row_index + 2, column=col_index + 1, value=self.categories[row_index])
                    else:
                        sheet.cell(row=row_index + 2, column=col_index + 1, value=auroc_px_ls[row_index] * 100)
                        sheet.cell(row=row_index + 2, column=col_index + 2, value=f1_px_ls[row_index] * 100)
                        sheet.cell(row=row_index + 2, column=col_index + 3, value=ap_px_ls[row_index] * 100)
                        sheet.cell(row=row_index + 2, column=col_index + 4, value=aupro_ls[row_index] * 100)
                        sheet.cell(row=row_index + 2, column=col_index + 5, value=auroc_sp_ls[row_index] * 100)
                        sheet.cell(row=row_index + 2, column=col_index + 6, value=f1_sp_ls[row_index] * 100)
                        sheet.cell(row=row_index + 2, column=col_index + 7, value=ap_sp_ls[row_index] * 100)
                    if row_index == len(self.categories) - 1:
                        if col_index == 0:
                            sheet.cell(row=row_index + 3, column=col_index + 1, value='mean')
                        else:
                            sheet.cell(row=row_index + 3, column=col_index + 1, value=auroc_px_mean * 100)
                            sheet.cell(row=row_index + 3, column=col_index + 2, value=f1_px_mean * 100)
                            sheet.cell(row=row_index + 3, column=col_index + 3, value=ap_px_mean * 100)
                            sheet.cell(row=row_index + 3, column=col_index + 4, value=aupro_mean * 100)
                            sheet.cell(row=row_index + 3, column=col_index + 5, value=auroc_sp_mean * 100)
                            sheet.cell(row=row_index + 3, column=col_index + 6, value=f1_sp_mean * 100)
                            sheet.cell(row=row_index + 3, column=col_index + 7, value=ap_sp_mean * 100)
            excel_file_path = os.path.join(self.output_dir, 'results.xlsx')
            csv_file_path = os.path.join(self.output_dir, 'results.csv')
            workbook.save(excel_file_path)
            df = pd.read_excel(excel_file_path)
            df.to_csv(csv_file_path, index=False)
